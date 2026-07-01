"""Load tabular uploads into pandas DataFrames for analysis.

Kept separate from the parsers (which target RAG/text) because analysis needs
typed DataFrames and, for Excel, access to raw formulas.
"""

from __future__ import annotations

import io

import pandas as pd

from audit_assistant.core.exceptions import ParsingError
from audit_assistant.domain.models import FileType


def load_dataframes(data: bytes, file_type: FileType) -> dict[str, pd.DataFrame]:
    """Return a mapping of sheet-name -> DataFrame (single entry for CSV)."""
    try:
        if file_type == FileType.CSV:
            try:
                return {"data": pd.read_csv(io.BytesIO(data))}
            except UnicodeDecodeError:
                return {"data": pd.read_csv(io.BytesIO(data), encoding="latin-1")}
        if file_type in {FileType.XLSX, FileType.XLS}:
            engine = "xlrd" if file_type == FileType.XLS else "openpyxl"
            return pd.read_excel(io.BytesIO(data), sheet_name=None, engine=engine)
    except Exception as exc:  # noqa: BLE001
        raise ParsingError(f"Could not load tabular data: {exc}") from exc
    raise ParsingError(f"{file_type.value} is not a tabular format.")


def count_formula_cells(data: bytes) -> dict[str, int]:
    """Count formula cells per sheet in an .xlsx file (empty for other types)."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=False)
    except Exception:  # noqa: BLE001 - not xlsx or unreadable
        return {}
    counts: dict[str, int] = {}
    for ws in wb.worksheets:
        n = 0
        for row in ws.iter_rows():
            for cell in row:
                if cell.data_type == "f" or (
                    isinstance(cell.value, str) and cell.value.startswith("=")
                ):
                    n += 1
        counts[ws.title] = n
    wb.close()
    return counts
